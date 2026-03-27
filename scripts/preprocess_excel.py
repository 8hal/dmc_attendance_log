#!/usr/bin/env python3
"""
동마클 최고 기록.xlsx 전처리 스크립트

- 정회원명단 시트: 닉네임 → 실명 매핑
- 회원 대회기록 시트: 전체 기록 정제 및 JSON 출력
"""

import json
import re
import sys
from datetime import datetime, date, time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl 없음. pip3 install openpyxl")
    sys.exit(1)

EXCEL_PATH = Path.home() / "Downloads" / "동마클 최고 기록.xlsx"
OUT_DIR = Path(__file__).parent.parent / "data"
OUT_DIR.mkdir(exist_ok=True)

# ── 종목명 정규화 ──────────────────────────────────────────
DISTANCE_MAP = {
    "full": "full",
    "half": "half",
    "10km": "10km",
    "하프": "half",
    "10km": "10km",
    "10k": "10km",
    "10km": "10km",
    "10.mile": "10km",
    "10kM": "10km",
    "42k": "full",
    "42km": "full",
    "21k": "half",
    "21km": "half",
    "5km": "5km",
    "5k": "5km",
    "100km": "ultra",
    "100k": "ultra",
    "100m": "ultra",
    "100M": "ultra",
    "70km": "ultra",
    "70k": "ultra",
    "50k": "ultra",
    "53km": "ultra",
}

def normalize_distance(raw):
    if raw is None:
        return None
    s = str(raw).strip().lower().replace("\u200b", "").replace("\t", "").replace(" ", "")
    if s in ("", "종목", "종목(full,half,10km)"):
        return None
    return DISTANCE_MAP.get(s, s)


# ── 기록 → "HH:MM:SS" 변환 ────────────────────────────────
def time_to_str(t):
    if isinstance(t, time):
        return t.strftime("%H:%M:%S")
    if isinstance(t, str):
        s = t.strip()
        if re.match(r"^\d{1,2}:\d{2}:\d{2}$", s):
            return s
    return None


# ── 날짜 변환 ─────────────────────────────────────────────
def date_to_str(d):
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d")
    if isinstance(d, date):
        return d.strftime("%Y-%m-%d")
    if isinstance(d, str):
        return d[:10]
    return None


# ── 닉네임 정규화 ─────────────────────────────────────────
def normalize_nick(raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    # openpyxl이 숫자 닉네임을 float으로 읽는 경우 (예: 123.0 → "123")
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


# ──────────────────────────────────────────────────────────
def main():
    print(f"📂 엑셀 로딩: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # 1. 닉 → 실명 매핑 (정회원명단)
    ws_members = wb["정회원명단"]
    nick_to_real = {}
    members_out = []
    all_nicks = []  # fuzzy 매칭용
    for row in ws_members.iter_rows(min_row=1, values_only=True):
        num = row[0]
        if not isinstance(num, (int, float)):
            continue
        nick = normalize_nick(row[2])
        realname = str(row[3]).strip() if row[3] else ""
        joindate = date_to_str(row[4])
        note = str(row[5]).strip() if row[5] else ""
        if nick and realname:
            nick_to_real[nick] = realname
            all_nicks.append(nick)
        members_out.append({
            "no": int(num),
            "nickName": nick,
            "realName": realname,
            "joinDate": joindate,
            "note": note,
        })

    # fuzzy 매핑 테이블: 대소문자 무시 + "포함" 매칭
    # "Josh" → "♥가족2_Josh", "james" → "James" 처리
    fuzzy_map = {}
    for nick in all_nicks:
        fuzzy_map[nick.lower()] = nick  # 소문자 → 원본 닉

    def resolve_nick(raw_nick):
        """닉네임 → 정회원명단 매핑. exact → lowercase → contains 순 시도"""
        if raw_nick in nick_to_real:
            return nick_to_real[raw_nick]
        lower = raw_nick.lower()
        if lower in fuzzy_map:
            return nick_to_real[fuzzy_map[lower]]
        # 정회원 닉에 raw_nick이 포함되는 경우 (예: "Josh" in "♥가족2_Josh")
        for nick in all_nicks:
            if lower in nick.lower() or nick.lower() in lower:
                return nick_to_real[nick]
        return None

    print(f"✅ 회원명단: {len(members_out)}명, 닉→실명 매핑 {len(nick_to_real)}건")

    # 2. 대회기록 파싱 (회원 대회기록)
    ws_records = wb["회원 대회기록"]
    records_out = []
    skipped = {"no_nick_match": [], "no_distance": [], "no_time": []}

    for i, row in enumerate(ws_records.iter_rows(min_row=2, values_only=True)):
        evt_date = date_to_str(row[0])
        race_name = str(row[1]).strip() if row[1] else None
        nick_raw = normalize_nick(row[2])
        distance_raw = row[4]
        time_raw = row[6]

        if not evt_date or not race_name or not nick_raw:
            continue
        if race_name in ("대회명", "종목"):
            continue

        realname = resolve_nick(nick_raw)
        if not realname:
            skipped["no_nick_match"].append(nick_raw)
            continue

        distance = normalize_distance(distance_raw)
        if not distance:
            skipped["no_distance"].append(f"{nick_raw} / {race_name} / {distance_raw}")
            continue

        finish_time = time_to_str(time_raw)
        if not finish_time:
            skipped["no_time"].append(f"{nick_raw} / {race_name} / {time_raw}")
            continue

        records_out.append({
            "memberNickName": nick_raw,
            "memberRealName": realname,
            "eventDate": evt_date,
            "eventName": race_name,
            "distance": distance,
            "finishTime": finish_time,
            "status": "confirmed",
            "confirmSource": "excel_import",
            "source": "excel",
            "importedAt": datetime.now().isoformat(),
        })

    # 3. 중복 제거 (같은 회원 + 대회 + 종목은 첫 번째만 유지)
    seen = set()
    deduped = []
    for r in records_out:
        key = (r["memberRealName"], r["eventName"], r["distance"])
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    dup_count = len(records_out) - len(deduped)

    # 4. 출력
    members_path = OUT_DIR / "members_from_excel.json"
    records_path = OUT_DIR / "race_records_from_excel.json"

    with open(members_path, "w", encoding="utf-8") as f:
        json.dump(members_out, f, ensure_ascii=False, indent=2)

    with open(records_path, "w", encoding="utf-8") as f:
        json.dump(deduped, f, ensure_ascii=False, indent=2)

    # 5. 요약 출력
    print(f"\n📊 전처리 결과 요약")
    print(f"  기록 파싱: {len(records_out)}건 → 중복 제거 후 {len(deduped)}건 ({dup_count}건 중복 제거)")

    no_match_uniq = sorted(set(skipped["no_nick_match"]))
    print(f"\n⚠️  닉네임 매핑 실패: {len(skipped['no_nick_match'])}건")
    if no_match_uniq:
        print(f"   미매핑 닉네임 ({len(no_match_uniq)}개):", ", ".join(no_match_uniq[:20]))
        if len(no_match_uniq) > 20:
            print(f"   ... 외 {len(no_match_uniq)-20}개")

    print(f"\n⚠️  종목 정보 없음: {len(skipped['no_distance'])}건")
    print(f"⚠️  기록 정보 없음: {len(skipped['no_time'])}건")

    print(f"\n💾 출력 파일:")
    print(f"  {members_path}")
    print(f"  {records_path}")

    # 6. 종목별 / 대회별 통계
    from collections import Counter
    dist_cnt = Counter(r["distance"] for r in deduped)
    print(f"\n📈 종목별 기록 수:")
    for d, cnt in dist_cnt.most_common():
        print(f"   {d}: {cnt}건")

    race_cnt = Counter(r["eventName"] for r in deduped)
    print(f"\n🏅 대회별 기록 수 (상위 15개):")
    for race, cnt in race_cnt.most_common(15):
        print(f"   {race}: {cnt}건")


if __name__ == "__main__":
    main()
