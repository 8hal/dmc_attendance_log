#!/usr/bin/env python3
"""
정회원 명단 엑셀/CSV → cleaned JSON 전처리

지원 형식:
  엑셀: 첫 시트 또는 --sheet 이름, row2+: B=순번, C=구분, D=닉네임, E=실명
  CSV: 동일 열 배치 (A 빈칸, B=순번, C=구분, D=닉네임, E=실명). 전화번호·비고는 저장하지 않음.

사용:
  python3 scripts/preprocess-members-excel.py path/to/명단.xlsx
  python3 scripts/preprocess-members-excel.py path/to/명단.csv --out scripts/data/members-2026-08-31-cleaned.json
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path


def normalize_nick(raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


def strip_prefix(nick):
    if nick == "♥동동♥":
        return "동동"
    if nick == "♥다빈♥":
        return "다빈"
    matched = re.match(r"^(?:♥가족\d+_|★자매_)(.+)$", nick)
    if matched:
        return matched.group(1)
    return nick


def member_from_cells(num, gubun, raw_nick, real):
    if gubun and gubun != "정회원":
        return None
    raw_nick = normalize_nick(raw_nick)
    real = str(real).strip() if real else ""
    if not raw_nick or not real:
        return None
    return {
        "순번": int(num),
        "구분": "정회원",
        "nickname": strip_prefix(raw_nick),
        "realName": real,
        "원본닉네임": raw_nick,
    }


def parse_workbook(wb, sheet_name=None):
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    members = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        num = row[1] if len(row) > 1 else None
        if not isinstance(num, (int, float)):
            continue
        member = member_from_cells(
            num,
            str(row[2]).strip() if len(row) > 2 and row[2] else "",
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else "",
        )
        if member:
            members.append(member)
    return members


def parse_csv(path):
    members = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if len(row) < 5:
                continue
            num_raw = str(row[1]).strip() if len(row) > 1 else ""
            if not num_raw.isdigit():
                continue
            member = member_from_cells(
                int(num_raw),
                str(row[2]).strip() if row[2] else "",
                row[3] if len(row) > 3 else None,
                row[4] if len(row) > 4 else "",
            )
            if member:
                members.append(member)
    return members


def load_members(src, sheet_name=None):
    suffix = src.suffix.lower()
    if suffix == ".csv":
        return parse_csv(src), "csv"
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 없음. pip3 install openpyxl")
        sys.exit(1)
    wb = openpyxl.load_workbook(src, data_only=True)
    sheet = sheet_name or wb.sheetnames[0]
    return parse_workbook(wb, sheet_name), sheet


def main():
    parser = argparse.ArgumentParser(description="정회원 명단 엑셀/CSV → cleaned JSON")
    parser.add_argument("excel", type=Path, help="엑셀 또는 CSV 파일 경로")
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="출력 JSON (기본: scripts/data/members-<stem>-cleaned.json)",
    )
    parser.add_argument("--sheet", default=None, help="시트 이름 (엑셀만, 기본: 첫 시트)")
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"파일 없음: {args.excel}")
        sys.exit(1)

    out = args.out
    if out is None:
        stem = args.excel.stem.replace(" ", "_")
        out = Path(__file__).parent / "data" / f"{stem}-cleaned.json"

    out.parent.mkdir(parents=True, exist_ok=True)

    members, source_label = load_members(args.excel, args.sheet)

    with open(out, "w", encoding="utf-8") as f:
        json.dump(members, f, ensure_ascii=False, indent=2)

    print(f"소스: {source_label}")
    print(f"정회원: {len(members)}명")
    print(f"출력: {out}")
    if members:
        print(f"  첫: {members[0]['nickname']} ({members[0]['realName']})")
        print(f"  끝: {members[-1]['nickname']} ({members[-1]['realName']})")


if __name__ == "__main__":
    main()
